import os
# Cria uma nova pasta

os.mkdir("")

# Função onde ficará a query

def arquivo_vendas(x):
    query = f"""
    """
    return query




import pandas as pd
import pyodbc
from prettytable import PrettyTable
import openpyxl
# Configuração da conexão com o banco de dados SQL Server
server = ''
database = ''
username = ''
password = ''
driver = '{ODBC Driver 17 for SQL Server}'

cnxn = pyodbc.connect('DRIVER=' + driver + ';SERVER=' + server + ';DATABASE=' + database + ';UID=' + username + ';PWD=' + password)

cursor = cnxn.cursor()



#GERADOR DOS ARQUIVOS xlsx
for x in range(1,4):
    print(x)
    

    cursor.execute(arquivo_vendas(x))

    results = cursor.fetchall()

    df = pd.DataFrame(results)

    workbook = openpyxl.Workbook()

    worksheet = workbook.active

    # Escreva os cabeçalhos das colunas na primeira linha da planilha
    column_names = [column[0] for column in cursor.description]
    for i, col_name in enumerate(column_names):
        worksheet.cell(row=1, column=i+1, value=col_name)

    # Escreva cada linha de resultado na planilha
    for row_num, row_data in enumerate(results):
        for col_num, col_data in enumerate(row_data):
            worksheet.cell(row=row_num+2, column=col_num+1, value=col_data)
    
    
    # condição para ver onde o arquivo tem que ficar, gerando e alocando na pasta designada
    if (x == 1):
        workbook.save(f'NomeDoArq{x}.xlsx')
        os.rename(f"NomeDoArq{x}.xlsx", f"Pasta/NomeDoArq{x}.xlsx")
    elif(x == 2):
        workbook.save(f'NomeDoArq{x}.xlsx')
        os.rename(f"NomeDoArq{x}.xlsx", f"Pasta/NomeDoArq{x}.xlsx")
    elif(x == 3):
        workbook.save(f'NomeDoArq{x}.xlsx') 
        os.rename(f"NomeDoArq{x}.xlsx", f"Pasta/NomeDoArq{x}.xlsx") 


